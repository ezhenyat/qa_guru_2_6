import zipfile
import pytest
import os
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook
import shutil


@pytest.fixture
def delete_test_folder():
    yield
    shutil.rmtree(os.path.abspath('./test_files'))


def test_create_archive():
    os.mkdir(os.path.abspath('./test_files'))
    archive = zipfile.ZipFile('test_files/archive.zip', 'w')
    archive.write('resources/example.pdf', arcname='example.pdf')
    archive.write('resources/example.csv', arcname='example.csv')
    archive.write('resources/example.xlsx', arcname='example.xlsx')
    archive.close()


def test_assert_files_content(delete_test_folder):
    with zipfile.ZipFile(os.path.abspath('./test_files/archive.zip')) as myzip:
        with myzip.open('example.pdf', 'r') as pdf_file:
            pdf_file = PdfReader(pdf_file)
            pdf_text = pdf_file.pages[0].extract_text()
            assert 'A Simple PDF File' in pdf_text, 'В pdf файле не содержится нужный текст'

        myzip.extract('example.csv', path='./test_files')
        with open(os.path.abspath('./test_files/example.csv')) as csv_file:
            csv_file = csv.reader(csv_file)
            for line_number, line in enumerate(csv_file, 1):
                if line_number == 4:
                    assert '362' in line[1], 'В  csv файле не содержится нужный текст'

        with myzip.open('example.xlsx', 'r') as xlsx_file:
            xlsx_file = load_workbook(xlsx_file)
            sheet = xlsx_file.active
            assert 'January' in sheet.cell(row=2, column=2).value, 'В xlsx файле не содержится нужный текст'
